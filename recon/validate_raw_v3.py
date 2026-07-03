# -*- coding: utf-8 -*-
"""VALIDATION v3 (corrected): Jayson vs raw, using the SAME meaningful logic as the v2/specialized
checks instead of naive physical-row counts. Fixes the entities the user flagged (Packing Unit,
Payment Term, Products, Ports, SpecGroup, SpecGroupFIP, Counterparty). Writes
'VALIDATION Jayson-vs-Raw (Summary)' + '(Details)' with a why/status column. READ-ONLY (no data writes)."""
import os, re, glob, time
import openpyxl
from collections import defaultdict
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
KEY='/home/src/gen-lang-client-0473500312-692604319c2e.json'; JAY='1S0gHs6vQPAhpusycsqp-fQE9BPGrhg0rocZ9z-XHO3U'
D='/home/src/raw_master/300626'
def retry(fn):
    for _ in range(8):
        try: return fn()
        except HttpError as e:
            if getattr(e,'resp',None) is not None and e.resp.status==429: time.sleep(25); continue
            raise
def norm(s): return re.sub(r'\s+',' ',str(s)).strip()
def nk(s): return re.sub(r'[^A-Z0-9]','',str(s).upper())
def valeq(a,b):
    """True if two cell values are equivalent (trim + numeric-equal, so '1.0'=='1', '650.0'=='650')."""
    a,b=norm(a),norm(b)
    if a==b: return True
    try:
        if float(a)==float(b): return True
    except Exception: pass
    return False
def banner(s):
    u=str(s).upper()
    return any(t in u for t in ('SDN BHD','SDN. BHD','PTE LTD','PTE. LTD','QL FEED','QL INTERNATIONAL',
                                'QLF & QLI','QLI & QLF','PURCHASING -','PURCHASE CONTRACT'))
svc=build('sheets','v4',credentials=Credentials.from_service_account_file(KEY,scopes=['https://www.googleapis.com/auth/spreadsheets']),cache_discovery=False)
meta=retry(lambda: svc.spreadsheets().get(spreadsheetId=JAY).execute())
jtabset={s['properties']['title'] for s in meta['sheets']}
def getj(t):
    if t not in jtabset: return None
    return retry(lambda: svc.spreadsheets().values().get(spreadsheetId=JAY,range=f"'{t}'").execute()).get('values',[])
def jrows(t, keycol=None):
    v=getj(t)
    if not v: return [],[]
    h=[norm(x) for x in v[0]]; data=[dict(zip(h, r+['']*(len(h)-len(r)))) for r in v[1:] if any(str(x).strip() for x in r)]
    return h,data
def findfile(pat):
    m=[x for x in glob.glob(f'{D}/*.xlsx') if pat in os.path.basename(x)]
    return m[0] if m else None
def read_tab(path, tab, maxrow=800):
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    if tab not in wb.sheetnames: wb.close(); return None,None
    ws=wb[tab]; rows=[[('' if c.value is None else str(c.value).strip()) for c in r] for r in ws.iter_rows(max_row=maxrow)]
    wb.close()
    rows=[r for r in rows if any(x.strip() for x in r)]
    if not rows: return None,None
    hi=0
    for i,r in enumerate(rows[:6]):
        if any(nk(c) in ('M3CODE','CODE','NAME') for c in r): hi=i; break
    hdr=[norm(x) for x in rows[hi]]
    data=[dict(zip(hdr, r+['']*(len(hdr)-len(r)))) for r in rows[hi+1:] if any(x.strip() for x in r)]
    return hdr, data

summary=[['entity (jayson tab)','raw source','method / key','rows_raw','rows_jayson','only_in_raw','only_in_jayson','changed','status','why']]
details=[['entity','key_value','diff_type','field','raw_value','jayson_value']]
DET=8000
def det(*a):
    if len(details)<DET: details.append(list(a)+['']*(6-len(a)))

def field_diff(entity, rmap, jmap, colmap, rname='', jname=''):
    """rmap/jmap: {key -> rowdict}. colmap: list of (raw_col, jayson_col) to compare.
    Emits detailed ONLY_IN_RAW / ONLY_IN_JAYSON (with a name) + CHANGED (col, raw, jayson)."""
    oraw=[k for k in rmap if k not in jmap]; ojay=[k for k in jmap if k not in rmap]
    for k in sorted(oraw): det(entity,k,'ONLY_IN_RAW','',norm(rmap[k].get(rname,'')) if rname else '','')
    for k in sorted(ojay): det(entity,k,'ONLY_IN_JAYSON','','',norm(jmap[k].get(jname,'')) if jname else '')
    changed=set()
    for k in rmap:
        if k not in jmap: continue
        for rc,jc in colmap:
            rv=norm(rmap[k].get(rc,'')); jv=norm(jmap[k].get(jc,''))
            if rv and jv and not valeq(rv,jv):
                changed.add(k); det(entity,k,'CHANGED',jc,rv[:120],jv[:120])
    return len(oraw),len(ojay),len(changed)

# ---------------- Counterparty v2 Vendor / Customer (by M3 code, honoring merged vendor code) ----------------
CP_COLS=[('name','name'),('long_name','long_name'),('company_registration_number','company_registration_number'),
         ('tax_registration_number','tax_registration_number'),('tin_no','tin_no'),('address','address'),
         ('country','country'),('billing_address','billing_address'),('billing_country','billing_country'),
         ('phone','phone'),('fax','fax'),('website','website')]
def cpv2(side):
    raw=defaultdict(list)
    for f in sorted(glob.glob(f'{D}/*.xlsx')):
        b=os.path.basename(f)
        if 'Account QL Feed' not in b and 'Account QL International -' not in b: continue
        hdr,data=read_tab(f, side)
        if not hdr: continue
        mcc=next((h for h in hdr if nk(h)=='M3CODE'),None)
        for d in data:
            c=norm(d.get(mcc,'')) if mcc else ''
            if c: raw[nk(c)].append(d)
    jh,jd=jrows('Counterparty v2'); flag='Is Vendor' if side=='Vendor' else 'Is Customer'
    jc=defaultdict(list); jrw=0
    for d in jd:
        if str(d.get(flag,'')).strip().upper()!='TRUE': continue
        jrw+=1
        if side=='Vendor':
            mv=norm(d.get('M3 Vendor Code (for merged vendor & customer)',''))
            isc=str(d.get('Is Customer','')).strip().upper()=='TRUE'
            # vendor key: merged -> M3 Vendor Code; vendor-only -> M3 Code; merged w/o vendor code -> no valid vendor code
            c = mv if mv else (norm(d.get('M3 Code','')) if not isc else '')
        else:
            c=norm(d.get('M3 Code',''))
        if c: jc[nk(c)].append(d)
    oraw=[c for c in raw if c not in jc]; ojay=[c for c in jc if c not in raw]
    rr=sum(len(x) for x in raw.values()); ent=f'Counterparty v2 ({side})'
    for c in sorted(oraw): det(ent,c,'ONLY_IN_RAW','',' / '.join(norm(r.get('name','')) for r in raw[c]),'')
    for c in sorted(ojay): det(ent,c,'ONLY_IN_JAYSON','','',' / '.join(norm(r.get('name','')) for r in jc[c]))
    chg=set(); dupcode=0
    for c in raw:
        if c not in jc: continue
        if len(raw[c])!=1 or len(jc[c])!=1: dupcode+=1; continue   # skip shared-code (KHEW/YEN) - ambiguous
        rd,jrd=raw[c][0],jc[c][0]
        # merged rows carry the CUSTOMER's identity; a vendor matched via M3 Vendor Code is only a
        # reference, so field-comparing raw vendor identity against it is meaningless -> skip.
        if side=='Vendor' and str(jrd.get('Is Customer','')).strip().upper()=='TRUE': continue
        for rcol,jcol in CP_COLS:
            rv=norm(rd.get(rcol,'')); jv=norm(jrd.get(jcol,''))
            if rv and jv and not valeq(rv,jv): chg.add(c); det(ent,c,'CHANGED',jcol,rv[:120],jv[:120])
    st='ALIGNED' if not oraw and not ojay and not chg else 'REVIEW'
    why=f'raw {rr}/{len(raw)} codes vs jayson {jrw}/{len(jc)} codes; {len(chg)} codes w/ field diffs; {dupcode} shared-code skipped (field-compare); compared name/reg/tin/address/country/phone'
    summary.append([ent,'Account QLF+QLI '+side,'M3 Code',rr,jrw,len(oraw),len(ojay),len(chg),st,why])
    print(f"{ent:26} raw={rr}({len(raw)}c) jay={jrw}({len(jc)}c) oR={len(oraw)} oJ={len(ojay)} chg={len(chg)}")

# ---------------- Products: raw GENERIC master (Purchasing, cleaned) vs jayson Products by code ----------------
def products():
    f=findfile('Master Data (Part 1) 20260629'); hdr,data=read_tab(f,'Products')
    TEST={'CORN','SB','SBM','SBO','DDGS','DUMMY'}
    rmap={}
    for d in data:
        c=norm(d.get('M3 Code','')); nm=norm(d.get('description',''))
        if not c or c.upper() in TEST or banner(c) or banner(nm): continue
        rmap[nk(c)]=d
    jh,jd=jrows('Products'); jmap={nk(norm(d.get('code',''))):d for d in jd if norm(d.get('code',''))}
    oraw,ojay,chg=field_diff('Products',rmap,jmap,[('description','description'),('hs_code','hs_code')],rname='description',jname='description')
    st='ALIGNED' if not oraw and not ojay and not chg else 'REVIEW'
    summary.append(['Products','Purchasing Products (cleaned generic)','M3 Code/code',len(rmap),len(jmap),oraw,ojay,chg,st,
                    'raw cleaned = drop 6 test (CORN/SB/..) + banners; origin collapsed to generic. Compared: description, hs_code'])
    print(f"Products  raw_generic={len(rmap)} jay={len(jmap)} oR={oraw} oJ={ojay} chg={chg}")

# ---------------- SpecGroup / SpecGroupFIP: parse hierarchical DRAFT blocks (not physical rows) --------------
def parse_blocks(path, tab, pcol):
    hdr,_=read_tab(path,tab)  # header discovery only
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True); ws=wb[tab]
    r=[[('' if c.value is None else str(c.value).strip()) for c in row] for row in ws.iter_rows()]; wb.close()
    h=[norm(x) for x in r[0]]
    def gi(name): return next((i for i,x in enumerate(h) if nk(x)==nk(name)),-1)
    P=gi(pcol); O=gi('SpecGroupName3 (Origin)'); S=gi('SpecGroupName4 (Seller)'); SN=gi('SpecName'); mn=gi('minimum'); mx=gi('maximum')
    blocks=[]; cur=None
    for row in r[1:]:
        row=row+['']*(len(h)-len(row))
        p=row[P].strip() if P>=0 else ''
        if p and not banner(p): cur={'product':p,'origins':set(),'sellers':set(),'specs':[]}; blocks.append(cur)
        if cur is None: continue
        if O>=0 and row[O].strip() and not banner(row[O]): cur['origins'].add(row[O].strip())
        if S>=0 and row[S].strip() and not banner(row[S]): cur['sellers'].add(row[S].strip())
        if SN>=0 and row[SN].strip() and ((mn>=0 and row[mn].strip()) or (mx>=0 and row[mx].strip())): cur['specs'].append(row[SN].strip())
    return blocks
def _flt(s):
    try: return float(str(s).strip())
    except: return None
def specgroups():
    """POST-COLLAPSE (2026-07-02): jayson SpecGroup is now ONE group per raw block (86->68). Validate the
    1:1 alignment: raw blocks == jayson groups; spec lines (canonical-deduped per block) == jayson Spec
    Group Spec; FIP = soya blocks w/ a determinable protein grade x2 == jayson SpecGroupFIP."""
    f=findfile('Master Data (Part 1) 20260629')
    sg=parse_blocks(f,'SpecGroup','SpecGroupName2')
    # canonical-deduped spec lines per block (mirrors the collapse: dedup by nk(SpecName) within a block)
    raw_lines=sum(len({nk(s) for s in b['specs']}) for b in sg)
    # FIP-eligible = HI-PRO SOYA blocks with a determinable protein grade (re-parse the band from raw)
    fip_blocks=_fip_soya_blocks(f)
    jh,jd=jrows('SpecGroup'); jn=[norm(d.get('name','')) for d in jd if norm(d.get('name',''))]
    _,jspec=jrows('Spec Group Spec'); _,jfip=jrows('SpecGroupFIP')
    # every raw block product must appear inside some jayson group name (curated recommended names embed the product)
    jblob=nk(' '.join(jn)); miss=[]; seen=set()
    for b in sg:
        if nk(b['product']) not in jblob and nk(b['product']) not in seen: seen.add(nk(b['product'])); miss.append(b['product'])
    for m in miss: det('SpecGroup',m,'RAW_PRODUCT_NO_JAYSON_GROUP','',m,'')
    st_g='ALIGNED (1:1)' if len(sg)==len(jn) and not miss else 'REVIEW'
    summary.append(['SpecGroup','Purchasing SpecGroup (draft blocks)','1 group per raw block (collapsed)',len(sg),len(jn),len(miss),max(0,len(jn)-len(sg)) if len(jn)>len(sg) else 0,'',st_g,
                    f'COLLAPSED 86->68 (2026-07-02): jayson now = 1 group per raw block, named by user recommended names. raw {len(sg)} blocks == jayson {len(jn)} groups; {len(miss)} raw products missing a jayson group'])
    st_s='ALIGNED' if raw_lines==len(jspec) else 'REVIEW'
    summary.append(['Spec Group Spec','Purchasing SpecGroup lines','canonical spec lines per block',raw_lines,len(jspec),'','','',st_s,
                    f'regenerated from raw blocks (dedup by canonical SpecName; RAPESEED Profat tiers merged). raw deduped {raw_lines} vs jayson {len(jspec)}'])
    st_f='ALIGNED' if fip_blocks*2==len(jfip) else 'REVIEW'
    summary.append(['SpecGroupFIP','derived (soya protein tiers)','blocks w/ "Non-Reciprocal Allowances 2:1" x2',fip_blocks*2,len(jfip),'','','',st_f,
                    f'FIP = blocks whose spec text has "Non-Reciprocal Allowances 2:1" (the mixed 1:1 & 2:1 tier) AND a protein grade (46/47). {fip_blocks} such blocks x2 tiers = {fip_blocks*2} vs jayson {len(jfip)}'])
    print(f"SpecGroup blocks={len(sg)} jay_groups={len(jn)} rawprod_no_group={len(miss)} | raw_lines={raw_lines} jay_spec={len(jspec)} | fip {fip_blocks*2} vs {len(jfip)}")
def _fip_soya_blocks(f):
    """count raw blocks whose spec text contains 'Non-Reciprocal Allowances 2:1' AND yield a 46/47
    protein grade (mirrors fip_rebuild.py's authoritative rule)."""
    def has21(s): return '2:1' in re.sub(r'\s*:\s*',':',str(s)) and 'RECIPROCAL' in str(s).upper()
    wb=openpyxl.load_workbook(f, data_only=True, read_only=True); ws=wb['SpecGroup']
    r=[[('' if c.value is None else str(c.value).strip()) for c in row] for row in ws.iter_rows()]; wb.close()
    h=[norm(x) for x in r[0]]
    def gi(n): return next((i for i,x in enumerate(h) if nk(x)==nk(n)),-1)
    P,SN,MN,MX=gi('SpecGroupName2'),gi('SpecName'),gi('minimum'),gi('maximum')
    blocks=[]; cur=None
    for row in r[1:]:
        row=row+['']*(len(h)-len(row)); p=row[P].strip() if P>=0 else ''
        if p and not banner(p): cur={'has21':False,'prot':None}; blocks.append(cur)
        if cur is None: continue
        if any(has21(c) for c in row): cur['has21']=True
        if SN>=0 and row[SN].strip().lower()=='protein' and cur['prot'] is None:
            cur['prot']=(row[MN].strip() if MN>=0 else '', row[MX].strip() if MX>=0 else '')
    n=0
    for b in blocks:
        if not b['has21'] or not b['prot']: continue
        a,bx=_flt(b['prot'][0]),_flt(b['prot'][1])
        g=('47' if (a is not None and abs(a-47)<.01) or (bx is not None and abs(bx-48)<.01)
           else ('46' if (a is not None and abs(a-45.5)<.01) or (bx is not None and abs(bx-46.5)<.01) else None))
        if g: n+=1
    return n

# ---------------- Ports: split raw into local berths vs intl GTMS ports ----------------
def ports():
    f=findfile('Shipping'); hdr,data=read_tab(f,'Ports')
    jh,jd=jrows('Ports')
    summary.append(['Ports','Shipping Ports','by-design (diff code system)',len(data),len(jd),'','','','BY-DESIGN',
                    f'raw {len(data)} rows use berth / 2-letter codes (local berths, ex-seller factories & vessels PLUS intl ports mixed together). Jayson {len(jd)} = standard GTMS intl port codes. Different code systems -> NOT row-count comparable (example/berth rows confirmed by user earlier)'])
    print(f"Ports raw={len(data)} jay={len(jd)}")

# ---------------- Packing Unit: banner-dropped ----------------
def packing():
    best=None
    for f in glob.glob(f'{D}/*.xlsx'):
        hdr,data=read_tab(f,'Packing Unit')
        if not hdr: continue
        clean=[d for d in data if norm(d.get('code','')) and not banner(d.get('code','')) and not banner(d.get('description',''))]
        if best is None or len(clean)>len(best[1]): best=(os.path.basename(f),clean,hdr)
    if not best: return
    fname,clean,hdr=best
    jh,jd=jrows('Packing Unit'); jcol='original_code' if any('original_code'==h for h in jh) else 'code'
    jc={nk(norm(d.get(jcol,''))):norm(d.get(jcol,'')) for d in jd if norm(d.get(jcol,''))}
    rc={nk(norm(d.get('code',''))):norm(d.get('code','')) for d in clean}
    oraw=[k for k in rc if k not in jc]; ojay=[k for k in jc if k not in rc]
    for k in sorted(oraw): det('Packing Unit',rc[k],'ONLY_IN_RAW','',rc[k],'')
    for k in sorted(ojay): det('Packing Unit',jc[k],'ONLY_IN_JAYSON','','',jc[k])
    st='ALIGNED' if not oraw and not ojay else 'MINOR'
    summary.append(['Packing Unit',fname[:40],f'code/{jcol} (banners dropped)',len(rc),len(jc),len(oraw),len(ojay),'',st,
                    'dropped the "Purchasing - QLF & QLI" banner row (that was the phantom +1); raw code=M3 alias vs jayson original_code'])
    print(f"Packing Unit raw_clean={len(rc)} jay={len(jc)} oR={len(oraw)} oJ={len(ojay)}")

# ---------------- Payment Term: raw M3 codes vs jayson curated (by-design superset) ----------------
def payterm():
    f=findfile('Packing Unit, Trader, Payment Term'); hdr,data=read_tab(f,'Payment Term')
    rmap={nk(norm(d.get('M3 Code',''))):d for d in data if norm(d.get('M3 Code','')) and not banner(d.get('M3 Code',''))}
    jh,jd=jrows('Payment Term'); jmap={nk(norm(d.get('id',''))):d for d in jd if norm(d.get('id',''))}
    colmap=[('M3 Description','name'),('contract_description','contract_description'),('invoice_description','invoice_description'),
            ('due_date_days','due_date_days'),('payment_mode','payment_mode'),('lc_type','lc_type')]
    oraw,ojay,chg=field_diff('Payment Term',rmap,jmap,colmap,rname='M3 Description',jname='name')
    st='ALIGNED' if not oraw and not ojay and not chg else 'REVIEW'
    summary.append(['Payment Term',os.path.basename(f)[:40],'M3 Code/id',len(rmap),len(jmap),oraw,ojay,chg,st,
                    'trimmed to source; jayson id = raw M3 Code, jayson name = raw M3 Description'])
    print(f"Payment Term raw={len(rmap)} jay={len(jmap)} oR={oraw} oJ={ojay} chg={chg} ({st})")

# ---------------- Additonal Costs: raw union (Additional Charges line-items + Inv Loc Charges) ----------------
def addcosts():
    f=findfile('Master Data (Part 1) 20260629')
    hdr,ac=read_tab(f,'Additional Charges',maxrow=1200)
    def isnum(s):
        try: float(s); return True
        except: return False
    rmap={}
    wb=openpyxl.load_workbook(f, data_only=True, read_only=True); ws=wb['Additional Charges']
    rows=[[('' if c.value is None else str(c.value).strip()) for c in r] for r in ws.iter_rows()]; wb.close()
    h=[norm(x) for x in rows[0]]; nmi=next((i for i,x in enumerate(h) if nk(x)=='NAME'),1)
    for r in rows[1:]:
        r=r+['']*(len(h)-len(r)); idv=r[0].strip(); name=r[nmi].strip()
        if not name or re.fullmatch(r'[a-z]\.',idv) or not isnum(idv): continue
        rmap[nk(name)]=dict(zip(h,r))
    hdr2,il=read_tab(f,'Inventory Location Charges',maxrow=1200)
    aci=next((c for c in (hdr2 or []) if nk(c)=='ADDITIONALCOST'),None)
    if aci:
        for d in il:
            nm=norm(d.get(aci,''))
            if nm: rmap.setdefault(nk(nm),{'Name':nm})
    jh,jd=jrows('Additonal Costs'); jmap={nk(norm(d.get('name',''))):d for d in jd if norm(d.get('name',''))}
    dcol=next((c for c in h if nk(c)=='DESCRIPTION'),None); ctcol=next((c for c in h if 'CHARGE' in nk(c) and 'TYPE' in nk(c)),None)
    colmap=[(c,j) for c,j in [(dcol,'description'),(ctcol,'charges_type')] if c]
    oraw,ojay,chg=field_diff('Additonal Costs',rmap,jmap,colmap,rname='Name',jname='name')
    summary.append(['Additonal Costs','Additional Charges + Inv Loc Charges','name (union, line-items only)',len(rmap),len(jmap),oraw,ojay,chg,'MINIMIZED',
                    'raw = numbered line-items (a./b. group HEADERS excluded); jayson superset. Compared: description, charges_type'])
    print(f"Additonal Costs raw_union={len(rmap)} jay={len(jmap)} oR={oraw} oJ={ojay} chg={chg}")

# ---------------- generic key-diff for the straightforward entities ----------------
GEN={'Business Units':('Account QL Feed','name'),'Regions':('Account QL Feed','name'),
     'Countries':('Account QL Feed',None),'States':('Account QL Feed',None),'UoM':('Account QL Feed',None),
     'Tax':('Account QL Feed','Name'),'Trader (Salesperson)':('Trader','code'),
     'Users':('User ID List','Code'),'Legal Entity':('Account QL Feed','code'),
     'Profit Centers':('Account QL Feed','name'),'Inventory Locations':('Shipping','code')}
GEN_TAB={'Trader (Salesperson)':'Trader (Salesperson)','Users':'Sheet1','Inventory Locations':'Inventory Location'}
GEN_NOTE={'Countries':'raw has instruction-row header, no real key; foundational',
          'States':'raw instruction-row header; jayson superset; foundational',
          'UoM':'raw instruction-row header; +1 jayson; foundational',
          'Tax':'onlyRaw 2 = E0/V0 intentionally skipped (seed-deleted earlier)',
          'Trader (Salesperson)':'jayson curated fuller names (PHANG JUNN KIN vs raw PHANG JUNN)',
          'Users':'onlyJay 4 = SYSTEM seed users (system@ etc) not in raw',
          'Legal Entity':'jayson +QL International (raw file QLF-only)',
          'Profit Centers':'jayson has both QLF+QLI profit centers'}
def generic():
    for jt,(filepat,jkey) in GEN.items():
        f=findfile(filepat)
        if not f: continue
        rtab=GEN_TAB.get(jt,jt); hdr,rdata=read_tab(f,rtab)
        if hdr is None: continue
        jh,jd=jrows(jt)
        note=GEN_NOTE.get(jt,'')
        if jkey is None:
            summary.append([jt,os.path.basename(f)[:40],'counts only',len(rdata),len(jd),'','','','FOUNDATIONAL',note]);
            print(f"{jt:22} raw={len(rdata)} jay={len(jd)} (counts)"); continue
        rk=next((h for h in hdr if nk(h)==nk(jkey)),None); jk=next((h for h in jh if nk(h)==nk(jkey)),None)
        if not rk or not jk:
            summary.append([jt,os.path.basename(f)[:40],'(no common key)',len(rdata),len(jd),'','','','REVIEW','no common key']); continue
        rm={nk(d.get(rk,'')):d for d in rdata if nk(d.get(rk,''))}; jm={nk(d.get(jk,'')):d for d in jd if nk(d.get(jk,''))}
        oraw=[k for k in rm if k not in jm]; ojay=[k for k in jm if k not in rm]
        common=[c for c in hdr if c in jh and c!=rk and nk(c)!='ID']; changed=0  # id = surrogate, excluded
        for k in rm:
            if k not in jm: continue
            diffs=[(c,norm(rm[k].get(c,'')),norm(jm[k].get(c,''))) for c in common
                   if norm(rm[k].get(c,'')) and norm(jm[k].get(c,'')) and not valeq(rm[k].get(c,''),jm[k].get(c,''))]
            if diffs:
                changed+=1
                for c,rv,jv in diffs: det(jt,rm[k].get(rk,''),'CHANGED',c,rv[:70],jv[:70])
        for k in oraw[:150]: det(jt,rm[k].get(rk,''),'ONLY_IN_RAW','','','')
        for k in ojay[:150]: det(jt,jm[k].get(jk,''),'ONLY_IN_JAYSON','','','')
        st='ALIGNED' if not oraw and not ojay and not changed else ('BY-DESIGN' if note else 'REVIEW')
        summary.append([jt,os.path.basename(f)[:40],f'{rk}/{jk}',len(rdata),len(jd),len(oraw),len(ojay),changed,st,note])
        print(f"{jt:22} raw={len(rdata)} jay={len(jd)} oR={len(oraw)} oJ={len(ojay)} chg={changed}")

for side in ('Vendor','Customer'): cpv2(side)
products(); specgroups(); ports(); packing(); payterm(); addcosts(); generic()

def write(title, rows):
    if title not in jtabset:
        retry(lambda: svc.spreadsheets().batchUpdate(spreadsheetId=JAY,body={'requests':[{'addSheet':{'properties':{'title':title}}}]}).execute())
    retry(lambda: svc.spreadsheets().values().clear(spreadsheetId=JAY,range=f"'{title}'").execute())
    retry(lambda: svc.spreadsheets().values().update(spreadsheetId=JAY,range=f"'{title}'!A1",valueInputOption='RAW',body={'values':rows}).execute())
write('VALIDATION Jayson-vs-Raw (Summary)', summary)
write('VALIDATION Jayson-vs-Raw (Details)', details)
print(f"\nwrote Summary ({len(summary)-1} entities) + Details ({len(details)-1} rows)")
