# -*- coding: utf-8 -*-
"""VALIDATION (1): Jayson sheet vs ALL raw dept files. For each entity, pick the authoritative raw
source (max data rows among files that have that tab), diff by business key -> only_in_raw /
only_in_jayson / changed(field). Writes 'VALIDATION Jayson-vs-Raw (Summary)' + '(Details)'. Read-only."""
import os, re, glob, time
import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
KEY='/home/src/gen-lang-client-0473500312-692604319c2e.json'; JAY='1S0gHs6vQPAhpusycsqp-fQE9BPGrhg0rocZ9z-XHO3U'
D='/home/src/raw_master/300626'
# raw tab -> jayson tab (exact names match unless aliased here)
ALIAS={'Additional Charges':'Additonal Costs','Trader (Salesperson)':'Trader (Salesperson)',
       'Inventory Location':'Inventory Locations','Sheet1':'Users',
       'Inventory Locations_QLF_29.6.26':'Inventory Locations','Inventory Locations_QLI_29.6.26':'Inventory Locations',
       'Warehouse Charges':'Inventory Location Charges'}
SKIP_TABS={'Payment Method','PaymentMtd+PaymentTerm','Payment Term- swee ee'}
KEYPREF=['M3 Code','code','Code','name','Name']
# entities not key-comparable to raw -> counts + note only (no detail rows)
NOTES={
 'Ports':'raw uses berth/2-letter codes vs jayson GTMS codes - different systems (expected)',
 'SpecGroup':'raw SpecGroup is a hierarchical DRAFT (blocks, not tabular by key) - counts only',
 'SpecGroupFIP':'raw SpecGroupFIP is a hierarchical draft - counts only',
 'Countries':'raw has instruction-row header (no real key); foundational/stable',
 'States':'raw has instruction-row header (no real key); foundational/stable',
 'UoM':'raw has instruction-row header (no real key)',
 'Payment Term':'raw uses M3 codes as names; jayson uses curated full names - not name-joinable (reconciled in pass 1/2)',
}
KEY_ALIAS={'Packing Unit':('code','original_code')}  # (raw col, jayson col)
def retry(fn):
    for _ in range(8):
        try: return fn()
        except HttpError as e:
            if getattr(e,'resp',None) is not None and e.resp.status==429: time.sleep(25); continue
            raise
def norm(s): return re.sub(r'\s+',' ',str(s)).strip()
def nk(s): return re.sub(r'[^A-Z0-9]','',str(s).upper())
svc=build('sheets','v4',credentials=Credentials.from_service_account_file(KEY,scopes=['https://www.googleapis.com/auth/spreadsheets']),cache_discovery=False)
def getj(t):
    try: return retry(lambda: svc.spreadsheets().values().get(spreadsheetId=JAY,range=f"'{t}'").execute()).get('values',[])
    except Exception: return None
meta=retry(lambda: svc.spreadsheets().get(spreadsheetId=JAY).execute())
jtabs={s['properties']['title'] for s in meta['sheets']}

def read_raw(path, tab):
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    if tab not in wb.sheetnames: return None
    ws=wb[tab]; rows=[[('' if c.value is None else str(c.value).strip()) for c in r] for r in ws.iter_rows(min_row=1, max_row=400)]
    rows=[r for r in rows if any(x.strip() for x in r)]  # drop fully-blank rows
    if not rows: return None
    # find header row = first row (within first 6) containing a preferred key col
    hi=0
    for i,r in enumerate(rows[:6]):
        if any(any(nk(c)==nk(k) for c in r) for k in KEYPREF): hi=i; break
    hdr=[norm(x) for x in rows[hi]]
    data=[dict(zip(hdr, r)) for r in rows[hi+1:] if any(x.strip() for x in r)]
    return hdr, data

def common_key(jt, rhdr, jhdr):
    if jt in KEY_ALIAS:
        rc,jc=KEY_ALIAS[jt]
        r=next((h for h in rhdr if nk(h)==nk(rc)),None); j=next((h for h in jhdr if nk(h)==nk(jc)),None)
        if r and j: return r,j
    for k in KEYPREF:                       # pick a key present in BOTH sides (aligned join)
        r=next((h for h in rhdr if nk(h)==nk(k)),None); j=next((h for h in jhdr if nk(h)==nk(k)),None)
        if r and j: return r,j
    return None,None

# collect raw (jtab -> list of (file,tab,hdr,data)); choose max-rows source
sources={}
for f in sorted(glob.glob(f'{D}/*.xlsx')):
    wb=openpyxl.load_workbook(f, read_only=True);
    for tab in wb.sheetnames:
        if tab in SKIP_TABS: continue
        jt=ALIAS.get(tab, tab)
        if jt not in jtabs: continue
        rr=read_raw(f, tab)
        if not rr or not rr[0]: continue
        sources.setdefault(jt, []).append((os.path.basename(f), tab, rr[0], rr[1]))

summary=[['entity (jayson tab)','raw source file','raw tab','key','rows_raw','rows_jayson','only_in_raw','only_in_jayson','changed_rows','note']]
details=[['entity','key_value','diff_type','field','raw_value','jayson_value','raw_source']]
DET_CAP=1500

def cpv2_recon(side):
    """raw Vendor/Customer (Account QLF + QLI combined) vs Counterparty v2 Is Vendor/Is Customer, by M3 code.
    Vendor side also honors the merged 'M3 Vendor Code' column (vendor code lives there for V+C merged rows)."""
    raw_codes={}
    for f in sorted(glob.glob(f'{D}/*.xlsx')):
        if 'Account QL Feed' not in f and 'Account QL International -' not in f: continue
        rr=read_raw(f, side)
        if not rr or not rr[0]: continue
        hdr,data=rr
        mcc=next((h for h in hdr if nk(h)==nk('M3 Code')),None)
        nmc=next((h for h in hdr if nk(h) in (nk('name'),nk('Name'))),None)
        for d in data:
            code=norm(d.get(mcc,'')) if mcc else ''
            if not code: continue
            raw_codes.setdefault(code,[]).append(norm(d.get(nmc,'')) if nmc else '')
    jv=getj('Counterparty v2'); jh=[norm(x) for x in jv[0]]
    def ji(n): return next((i for i,c in enumerate(jh) if nk(c)==nk(n)),-1)
    i_mc,i_iv,i_ic,i_nm,i_vc=ji('M3 Code'),ji('Is Vendor'),ji('Is Customer'),ji('name'),ji('M3 Vendor Code (for merged vendor & customer)')
    flag=i_iv if side=='Vendor' else i_ic
    jrows=0; jcodes={}
    for r in jv[1:]:
        r=r+['']*(len(jh)-len(r))
        if r[flag].strip().upper()!='TRUE': continue
        jrows+=1
        code=r[i_mc].strip()
        if side=='Vendor' and i_vc>=0 and r[i_vc].strip(): code=r[i_vc].strip()
        if code: jcodes.setdefault(code,[]).append(r[i_nm].strip())
    only_raw=[c for c in raw_codes if c not in jcodes]; only_jay=[c for c in jcodes if c not in raw_codes]
    raw_rows=sum(len(x) for x in raw_codes.values())
    ent=f'Counterparty v2 ({side})'
    note=f'vs Counterparty v2 Is {side}; key=M3 Code; raw {raw_rows} rows/{len(raw_codes)} distinct codes; jayson {jrows} rows/{len(jcodes)} distinct codes'
    summary.append([ent,'Account QLF+QLI',side,'M3 Code',raw_rows,jrows,len(only_raw),len(only_jay),'',note])
    for c in sorted(only_raw)[:200]:
        if len(details)<DET_CAP: details.append([ent,c,'ONLY_IN_RAW','',' / '.join(raw_codes[c])[:80],'','Account QLF+QLI'])
    for c in sorted(only_jay)[:200]:
        if len(details)<DET_CAP: details.append([ent,c,'ONLY_IN_JAYSON','','',' / '.join(jcodes[c])[:80],'Counterparty v2'])
    print(f"{ent:28} raw={raw_rows:4}({len(raw_codes)} codes) jay={jrows:4}({len(jcodes)} codes) onlyRaw={len(only_raw):3} onlyJay={len(only_jay):3}")
for side in ('Vendor','Customer'):   # raw vendor/customer -> Counterparty v2 (Is Vendor/Is Customer)
    cpv2_recon(side)
for jt in sorted(sources):
    if jt in ('Vendor','Customer'): continue   # handled via cpv2_recon above
    cand=sources[jt]
    fname,rtab,rhdr,rdata=max(cand, key=lambda x: len(x[3]))  # authoritative = most rows
    jv=getj(jt)
    if not jv: continue
    jhdr=[norm(x) for x in jv[0]]; jdata=[dict(zip(jhdr, r+['']*(len(jhdr)-len(r)))) for r in jv[1:] if any(str(x).strip() for x in r)]
    if jt in NOTES:                          # structural: counts + note only
        summary.append([jt,fname,rtab,'-',len(rdata),len(jdata),'','','',NOTES[jt]]);
        print(f"{jt:28} raw={len(rdata):4} jay={len(jdata):4}  NOTE: {NOTES[jt][:44]}"); continue
    rk,jk=common_key(jt, rhdr, jhdr)
    if not rk or not jk:
        summary.append([jt,fname,rtab,'(no common key)',len(rdata),len(jdata),'','','','no common business key between raw & jayson'])
        continue
    rmap={nk(d.get(rk,'')):d for d in rdata if nk(d.get(rk,''))}
    jmap={nk(d.get(jk,'')):d for d in jdata if nk(d.get(jk,''))}
    common_cols=[h for h in rhdr if h in jhdr and h!=rk]
    only_raw=[k for k in rmap if k not in jmap]
    only_jay=[k for k in jmap if k not in rmap]
    changed=0
    for k in rmap:
        if k not in jmap: continue
        diffs=[]
        for col in common_cols:
            rv=norm(rmap[k].get(col,'')); jvv=norm(jmap[k].get(col,''))
            if rv!=jvv and not (rv=='' or jvv==''):  # ignore blank-vs-value (additive)
                diffs.append((col,rv,jvv))
        if diffs:
            changed+=1
            for col,rv,jvv in diffs:
                if len(details)<DET_CAP: details.append([jt, rmap[k].get(rk,''),'CHANGED',col,rv[:80],jvv[:80],fname])
    for k in only_raw[:200]:
        if len(details)<DET_CAP: details.append([jt, rmap[k].get(rk,''),'ONLY_IN_RAW','','', '',fname])
    for k in only_jay[:200]:
        if len(details)<DET_CAP: details.append([jt, jmap[k].get(jk,''),'ONLY_IN_JAYSON','','','',fname])
    summary.append([jt,fname,rtab,f"{rk}/{jk}",len(rdata),len(jdata),len(only_raw),len(only_jay),changed,''])
    print(f"{jt:28} raw={len(rdata):4} jay={len(jdata):4} onlyRaw={len(only_raw):3} onlyJay={len(only_jay):3} changed={changed:3}  key={rk}/{jk}")

def write(title, rows):
    if title not in [s['properties']['title'] for s in meta['sheets']]:
        retry(lambda: svc.spreadsheets().batchUpdate(spreadsheetId=JAY,body={'requests':[{'addSheet':{'properties':{'title':title}}}]}).execute())
    retry(lambda: svc.spreadsheets().values().clear(spreadsheetId=JAY,range=f"'{title}'").execute())
    retry(lambda: svc.spreadsheets().values().update(spreadsheetId=JAY,range=f"'{title}'!A1",valueInputOption='RAW',body={'values':rows}).execute())
write('VALIDATION Jayson-vs-Raw (Summary)', summary)
write('VALIDATION Jayson-vs-Raw (Details)', details)
print(f"\nwrote report tabs: Summary ({len(summary)-1} entities), Details ({len(details)-1} rows, capped {DET_CAP})")
