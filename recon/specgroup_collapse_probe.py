# -*- coding: utf-8 -*-
"""READ-ONLY probe of the 3 rebuild risks: (1) raw block SpecNames not resolvable to Specifications
(typo/gap set), (2) each SOYA block's Protein band -> FIP grade, (3) code->sales_desc consistency
carried from current SpecGroup via junction. No writes."""
import glob, re, time
from collections import defaultdict, Counter
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import openpyxl
KEY='/home/src/gen-lang-client-0473500312-692604319c2e.json'; JAY='1S0gHs6vQPAhpusycsqp-fQE9BPGrhg0rocZ9z-XHO3U'; D='/home/src/raw_master/300626'
def nk(s): return re.sub(r'[^A-Z0-9]','',str(s).upper())
def norm(s): return re.sub(r'\s+',' ',str(s)).strip()
def banner(s):
    u=str(s).upper(); return any(t in u for t in ('SDN BHD','PTE LTD','QL FEED','QL INTERNATIONAL','ALL SELLERS'))
def flt(s):
    try: return float(str(s).strip())
    except: return None
def retry(fn):
    for _ in range(8):
        try: return fn()
        except HttpError as e:
            if getattr(e,'resp',None) is not None and e.resp.status==429: time.sleep(25); continue
            raise
svc=build('sheets','v4',credentials=Credentials.from_service_account_file(KEY,scopes=['https://www.googleapis.com/auth/spreadsheets']),cache_discovery=False)
def getj(t): return retry(lambda: svc.spreadsheets().values().get(spreadsheetId=JAY,range=f"'{t}'").execute()).get('values',[])
def rows2dicts(v):
    h=[c.strip() for c in v[0]]; return h,[dict(zip(h,r+['']*(len(h)-len(r)))) for r in v[1:]]

# recommended names
vb=getj('VALIDATION SpecGroup Blocks'); vh=[c.strip() for c in vb[0]]; rc=vh.index('recommended name')
rec={int(r[0]):r[rc].strip() for r in vb[1:] if r and r[0].strip().isdigit()}

# parse blocks with full specs
f=[x for x in glob.glob(f'{D}/*.xlsx') if 'Master Data (Part 1) 20260629' in x][0]
wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb['SpecGroup']
r=[[('' if c.value is None else str(c.value).strip()) for c in row] for row in ws.iter_rows()]
h=[norm(x) for x in r[0]]
def gi(n): return next((i for i,x in enumerate(h) if nk(x)==nk(n)),-1)
P,SN,mn,mx=gi('SpecGroupName2'),gi('SpecName'),gi('minimum'),gi('maximum')
blocks=[]; cur=None
for row in r[1:]:
    row=row+['']*(len(h)-len(row)); p=row[P].strip()
    if p and not banner(p): cur={'product':norm(p),'specs':[]}; blocks.append(cur)
    if cur is None: continue
    if SN>=0 and row[SN].strip() and ((mn>=0 and row[mn].strip()) or (mx>=0 and row[mx].strip())):
        cur['specs'].append((norm(row[SN]),row[mn].strip() if mn>=0 else '',row[mx].strip() if mx>=0 else ''))

# (1) SpecName resolution vs Specifications
spv=getj('Specifications'); sph,spd=rows2dicts(spv)
spec_canon={}   # nk -> canonical name
for d in spd:
    nm=d.get('name','').strip()
    if nm: spec_canon[nk(nm)]=nm
raw_specnames=Counter(); unresolved=Counter()
for b in blocks:
    for sn,_,_ in b['specs']:
        raw_specnames[sn]+=1
        if nk(sn) not in spec_canon: unresolved[sn]+=1
print(f"Specifications sheet: {len(spec_canon)} specs")
print(f"raw block distinct SpecNames: {len(raw_specnames)} | UNRESOLVED (not in Specifications by nk): {len(unresolved)}")
for sn,c in unresolved.most_common(): print(f"   UNRESOLVED '{sn}' x{c}  (nk={nk(sn)})")

# (2) SOYA blocks -> protein band -> grade
def grade(mn_,mx_):
    a,b=flt(mn_),flt(mx_)
    if a is not None and abs(a-47)<0.01: return '47'
    if b is not None and abs(b-48)<0.01: return '47'
    if (a is not None and abs(a-45.5)<0.01) or (b is not None and abs(b-46.5)<0.01): return '46'
    return None
print("\nSOYA blocks (FIP candidates):")
nfip=0
for i,b in enumerate(blocks,1):
    nm=rec.get(i,'')
    if 'SOYA' not in nm.upper(): continue
    prot=next(((mn_,mx_) for sn,mn_,mx_ in b['specs'] if sn.strip().lower()=='protein'),None)
    g=grade(*prot) if prot else None
    if g: nfip+=1
    print(f"   blk {i:>2} grade={g}  protein={prot}  {nm[:55]}")
print(f"SOYA blocks getting FIP (grade found) = {nfip} -> {nfip*2} FIP rows")

# (3) code -> sales_desc consistency
sgv=getj('SpecGroup'); sgh,sgd=rows2dicts(sgv)
g_sales={norm(d['name']):d.get('sales_spec_group_description','').strip() for d in sgd}
g_desc={norm(d['name']):d.get('description','').strip() for d in sgd}
sxv=getj('Spec Group x Product'); sxh,sxd=rows2dicts(sxv)
code_sales=defaultdict(Counter); code_desc=defaultdict(Counter)
for d in sxd:
    g=norm(d.get('spec_group','')); c=d.get('code','').strip()
    if not g or not c: continue
    if g_sales.get(g): code_sales[c][g_sales[g]]+=1
    if g_desc.get(g): code_desc[c][g_desc[g]]+=1
print("\ncode -> sales_desc (distinct values per code; >1 = conflict):")
for c in sorted(code_sales):
    vals=code_sales[c]
    flag='CONFLICT' if len(vals)>1 else 'ok'
    print(f"   {c:14} [{flag}] {len(vals)} distinct; top='{vals.most_common(1)[0][0][:45]}'")
# which block codes will have NO sales_desc
blockcodes=set()
wb2=openpyxl.load_workbook(f,read_only=True,data_only=True); ws2=wb2['SpecGroup']
MC=gi('M3 Code'); cur=None
for row in [[('' if c.value is None else str(c.value).strip()) for c in rr] for rr in ws2.iter_rows()][1:]:
    row=row+['']*(len(h)-len(row))
    if row[P].strip() and not banner(row[P].strip()): cur=True
    if cur and MC>=0 and row[MC].strip(): blockcodes.add(row[MC].strip())
print(f"\nblock codes with NO carried sales_desc: {sorted(blockcodes - set(code_sales))}")
print("PROBE DONE (read-only).")
