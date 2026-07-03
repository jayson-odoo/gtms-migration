# -*- coding: utf-8 -*-
"""READ-ONLY. Investigate FIP over-generation: dump raw Purchasing 'SpecGroupFIP' tab (rows, fip values,
distinct groups), the pre-collapse jayson backup (48), and my new 36, to find which specs GENUINELY have
a mixed 1:1 & 2:1 (fip1+fip2) tier vs where I wrongly gave every graded soya block FIP."""
import glob, re, csv, os
from collections import defaultdict, Counter
import openpyxl
D='/home/src/raw_master/300626'; BK='/home/src/recon/backup'
def norm(s): return re.sub(r'\s+',' ',str(s)).strip()
def nk(s): return re.sub(r'[^A-Z0-9]','',str(s).upper())
f=[x for x in glob.glob(f'{D}/*.xlsx') if 'Master Data (Part 1) 20260629' in x][0]
wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
print("raw Part1 tabs:", [t for t in wb.sheetnames if 'FIP' in t.upper() or 'SPEC' in t.upper()])
if 'SpecGroupFIP' in wb.sheetnames:
    ws=wb['SpecGroupFIP']
    rows=[[('' if c.value is None else str(c.value).strip()) for c in r] for r in ws.iter_rows()]
    rows=[r for r in rows if any(x.strip() for x in r)]
    h=[norm(x) for x in rows[0]]
    print("\nraw SpecGroupFIP header:", h)
    print("raw SpecGroupFIP data rows:", len(rows)-1)
    def gi(n): return next((i for i,x in enumerate(h) if nk(x)==nk(n)),-1)
    gN,sN,fp=gi('SpecGroupName'),gi('SpecName'),gi('fip')
    fipcol=Counter(); bygroup=defaultdict(list)
    for r in rows[1:]:
        r=r+['']*(len(h)-len(r))
        fv=r[fp].strip() if fp>=0 else ''
        fipcol[fv]+=1
        if gN>=0: bygroup[norm(r[gN])].append((norm(r[sN]) if sN>=0 else '', fv))
    print("raw fip value counts:", dict(fipcol))
    print(f"raw distinct SpecGroupName in FIP tab: {len(bygroup)}")
    for g,items in bygroup.items():
        fips=sorted({x[1] for x in items})
        print(f"   '{g[:55]}' specs={sorted({x[0] for x in items})} fips={fips} n={len(items)}")
else:
    print("no raw SpecGroupFIP tab")
wb.close()

# pre-collapse jayson backup
p=f"{BK}/SpecGroupFIP.pre-collapse.csv"
if os.path.exists(p):
    with open(p) as fh: rr=list(csv.reader(fh))
    h=rr[0]; print(f"\npre-collapse jayson SpecGroupFIP backup: {len(rr)-1} rows, header={h}")
    gi2=lambda n: h.index(n) if n in h else -1
    gN,fp=gi2('SpecGroupName'),gi2('fip')
    bg=defaultdict(set)
    for r in rr[1:]:
        if len(r)>max(gN,fp): bg[norm(r[gN])].add(r[fp].strip())
    print(f"  distinct groups={len(bg)}; fip-value counts:", Counter(r[fp].strip() for r in rr[1:] if len(r)>fp))
