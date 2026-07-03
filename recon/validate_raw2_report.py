# -*- coding: utf-8 -*-
"""Consolidated corrected raw-vs-jayson report for user-flagged entities -> tab 'VALIDATION Jayson-vs-Raw v2'."""
import io, re, glob, time
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.errors import HttpError
import openpyxl
KEY='/home/src/gen-lang-client-0473500312-692604319c2e.json'; JAY='1S0gHs6vQPAhpusycsqp-fQE9BPGrhg0rocZ9z-XHO3U'
PUR='1SbnbbPgFkHHIy_5XnAm0ewVXX9Qw1vEa'; PTA='1H2h-wCSjCvRzcaefBye1zNHtBOOA5m2U'; PTB='1RNaMj4IbIiyAlGOOfBKtt39WCSy4pp27'
DIR='/home/src/raw_master/300626'
creds=Credentials.from_service_account_file(KEY,scopes=['https://www.googleapis.com/auth/drive.readonly','https://www.googleapis.com/auth/spreadsheets'])
drive=build('drive','v3',credentials=creds,cache_discovery=False); sheets=build('sheets','v4',credentials=creds,cache_discovery=False)
def nk(s): return re.sub(r'[^A-Z0-9]','',str(s).upper())
def norm(s): return re.sub(r'\s+',' ',str(s)).strip()
def isnum(s):
    try: float(s); return True
    except: return False
def banner(s): u=str(s).upper(); return 'SDN BHD' in u or 'PTE LTD' in u or 'QL FEED' in u or 'QL INTERNATIONAL' in u
def load(fid):
    buf=io.BytesIO(); dl=MediaIoBaseDownload(buf, drive.files().get_media(fileId=fid)); d=False
    while not d: _,d=dl.next_chunk()
    buf.seek(0); return openpyxl.load_workbook(buf, data_only=True, read_only=True)
def rows_wb(wb,tab):
    if tab not in wb.sheetnames: return []
    ws=wb[tab]; return [[('' if c.value is None else str(c.value).strip()) for c in r] for r in ws.iter_rows()]
def getj(t): return sheets.spreadsheets().values().get(spreadsheetId=JAY,range=f"'{t}'").execute().get('values',[])
pur=load(PUR)
rep=[['entity','raw_source','raw_count','jayson_count','only_in_raw','only_in_jayson','finding']]

# 1) ADDITONAL COSTS combined
ac=rows_wb(pur,'Additional Charges'); ach=[norm(x) for x in ac[0]]; nmi=ach.index('Name')
rawac={}; 
for r in ac[1:]:
    r=r+['']*(len(ach)-len(r)); idv=r[0].strip(); name=r[nmi].strip()
    if name and not re.fullmatch(r'[a-z]\.',idv) and isnum(idv): rawac[nk(name)]=name
il=rows_wb(pur,'Inventory Location Charges'); ilh=[norm(x) for x in il[0]]; aci=ilh.index('Additional Cost')
for r in il[1:]:
    r=r+['']*(len(ilh)-len(r)); name=r[aci].strip()
    if name and not banner(name): rawac.setdefault(nk(name),name)
jac=getj('Additonal Costs'); jni=[norm(x) for x in jac[0]].index('name')
jacset={nk(r[jni]) for r in jac[1:] if len(r)>jni and r[jni].strip()}
oraw=sorted(set(rawac)-jacset); ojay=len(jacset-set(rawac))
rep.append(['Additonal Costs','AdditionalCharges + InvLocCharges',len(rawac),len(jacset),len(oraw),ojay,
            'raw-only charges (mostly Inv Loc storage/warehouse tiers) to ADD: '+", ".join(rawac[k] for k in oraw[:18])])

# 2) PAYMENT TERM combined
def ptcodes(fid):
    r=rows_wb(load(fid),'Payment Term'); h=[norm(x) for x in r[0]]; ci=h.index('M3 Code')
    return {nk(x[ci]):x[ci].strip() for x in r[1:] if len(x)>ci and x[ci].strip()}
rawpt={}; rawpt.update(ptcodes(PTA)); rawpt.update(ptcodes(PTB))
jpt=getj('Payment Term'); jph=[norm(x) for x in jpt[0]]; jinv=jph.index('invoice_description'); jnm=jph.index('name')
jptk=set()
for r in jpt[1:]:
    r=r+['']*(len(jph)-len(r)); jptk.add(nk(r[jinv]) if r[jinv].strip() else nk(r[jnm]))
opt=sorted(set(rawpt)-jptk)
rep.append(['Payment Term','PT_A + PT_B (by M3 Code)',len(rawpt),len(jptk),len(opt),len(jptk-set(rawpt)),
            'raw codes not matched in jayson (by invoice_desc): '+", ".join(rawpt[k] for k in opt)+' | jayson richer/curated'])

# 3) COUNTERPARTY: raw Vendor+Customer union across ALL local raw files, by M3 code
rawcp=set()
for f in glob.glob(f'{DIR}/*.xlsx'):
    wb=openpyxl.load_workbook(f, read_only=True, data_only=True)
    for tab in ('Vendor','Customer'):
        rr=rows_wb(wb,tab)
        if not rr: continue
        hi=0
        for i,row in enumerate(rr[:3]):
            if any('M3' in str(c).upper() and 'CODE' in str(c).upper() for c in row): hi=i; break
        h=[norm(x) for x in rr[hi]]; ci=next((i for i,x in enumerate(h) if 'M3' in x.upper() and 'CODE' in x.upper()),None)
        if ci is None: continue
        for x in rr[hi+1:]:
            if len(x)>ci and x[ci].strip() and not banner(x[ci]): rawcp.add(nk(x[ci]))
jcp=getj('Counterparty v2'); jch=[norm(x) for x in jcp[0]]; mci=jch.index('M3 Code'); mvi=jch.index('M3 Vendor Code (for merged vendor & customer)')
jcpc=set()
for r in jcp[1:]:
    r=r+['']*(len(jch)-len(r))
    if r[mci].strip(): jcpc.add(nk(r[mci]))
    if r[mvi].strip(): jcpc.add(nk(r[mvi]))
oc=sorted(rawcp-jcpc)
rep.append(['Counterparty v2','raw Vendor+Customer (all files, M3 code)',len(rawcp),len(jcpc),len(oc),len(jcpc-rawcp),
            'raw M3 codes NOT in Cpv2 (missing counterparties): '+(", ".join(oc[:25]) if oc else 'none')])

# 4/5/6) SPEC GROUP / SPEC GROUP SPEC / FIP structural
def parse_blocks(tab,pcol):
    r=rows_wb(pur,tab); h=[norm(x) for x in r[0]]
    P=h.index(pcol); O=h.index('SpecGroupName3 (Origin)'); S=h.index('SpecGroupName4 (Seller)'); SN=h.index('SpecName'); mn=h.index('minimum'); mx=h.index('maximum')
    blocks=[]; cur=None
    for row in r[1:]:
        row=row+['']*(len(h)-len(row)); p=row[P].strip()
        if p and not banner(p): cur={'product':p,'specs':[]}; blocks.append(cur)
        if cur is None: continue
        if row[SN].strip() and (row[mn].strip() or row[mx].strip()): cur['specs'].append(row[SN].strip())
    return blocks
sg=parse_blocks('SpecGroup','SpecGroupName2'); fip=parse_blocks('SpecGroupFIP','SpecGroupName')
jsg=getj('SpecGroup'); jsgn=[r[[x.strip() for x in jsg[0]].index('name')].strip() for r in jsg[1:] if len(r)>1 and r[1].strip()]
jspec=len(getj('Spec Group Spec'))-1; jfip=len(getj('SpecGroupFIP'))-1
jblob=nk(' '.join(jsgn))
sgmiss=sorted({b['product'] for b in sg if nk(b['product']) not in jblob and not b['product'].startswith('PMQ')})
rep.append(['SpecGroup (groups)','SpecGroup blocks',len(sg),len(jsgn),len(sgmiss),'-',
            f'{len(sg)} raw blocks/{len(set(nk(b["product"]) for b in sg))} products vs {len(jsgn)} jayson groups. Raw products w/o jayson group: '+", ".join(sgmiss)])
rep.append(['Spec Group Spec','SpecGroup spec lines',sum(len(b['specs']) for b in sg),jspec,'-','-',
            f'raw {sum(len(b["specs"]) for b in sg)} spec lines vs jayson {jspec} (jayson expanded per group)'])
rep.append(['SpecGroupFIP','SpecGroupFIP spec lines',sum(len(b['specs']) for b in fip),jfip,'-','-',
            f'RAW FIP={sum(len(b["specs"]) for b in fip)} lines / {len(set(nk(b["product"]) for b in fip))} products with FULL spec detail (Protein tiers+Moisture+Fibre+Fat+Aflatoxin+...); JAYSON FIP={jfip} = only Protein 2-tier for soya. UNDER-POPULATED.'])

def write(title,rows):
    meta=sheets.spreadsheets().get(spreadsheetId=JAY).execute()
    if title not in [s['properties']['title'] for s in meta['sheets']]:
        sheets.spreadsheets().batchUpdate(spreadsheetId=JAY,body={'requests':[{'addSheet':{'properties':{'title':title}}}]}).execute()
    sheets.spreadsheets().values().clear(spreadsheetId=JAY,range=f"'{title}'").execute()
    sheets.spreadsheets().values().update(spreadsheetId=JAY,range=f"'{title}'!A1",valueInputOption='RAW',body={'values':rows}).execute()
write('VALIDATION Jayson-vs-Raw v2', rep)
for r in rep[1:]: print(f"{r[0]:22} raw={r[2]} jay={r[3]} onlyRaw={r[4]} onlyJay={r[5]}\n   {r[6][:150]}")
print("\nwrote tab 'VALIDATION Jayson-vs-Raw v2'")
